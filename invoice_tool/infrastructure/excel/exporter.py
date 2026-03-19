from pathlib import Path
from typing import List, Union

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from invoice_tool.domain.models import InvoiceRecord


EXPORT_COLUMNS = [
    "文件名",
    "发票号码",
    "开票日期",
    "购买方名称",
    "购买方税号",
    "销售方名称",
    "销售方税号",
    "金额",
    "税额",
    "价税合计",
]

INCOMPLETE_ROW_FILL = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")


def export_records_to_excel(records: List[InvoiceRecord], output_path: Union[str, Path]) -> Path:
    target_path = Path(output_path)
    target_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(EXPORT_COLUMNS)

    for record in records:
        worksheet.append([getattr(record, column_name) for column_name in EXPORT_COLUMNS])

    _style_incomplete_rows(worksheet, records)
    workbook.save(target_path)
    return target_path


def _style_incomplete_rows(worksheet, records: List[InvoiceRecord]) -> None:
    for row_index, record in enumerate(records, start=2):
        if record.是否完整:
            continue
        for cell in worksheet[row_index]:
            cell.fill = INCOMPLETE_ROW_FILL
